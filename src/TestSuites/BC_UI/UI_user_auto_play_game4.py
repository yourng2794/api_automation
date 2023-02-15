#!/usr/bin/python
# -*- coding: UTF-8 -*-
import logging, sys, os, inspect, shutil, locale
from datetime import datetime
import time
from base64 import b64decode
from BaseTestCase import BaseCase
from O2oUtility import WriteLog, WriteDebugLog, IniValue, ApiCodeDetect
from O2oUtility import Compare_and_Click
from TestRun import TestRunning

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl.styles import Alignment

locale.setlocale(locale.LC_CTYPE, 'chinese')
currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
tooldir = os.path.dirname(os.path.dirname(currentdir))
if tooldir not in sys.path:
    sys.path.insert(0, tooldir)
iniFile = tooldir+'\\ini\\setting.ini'

class TestCase(BaseCase):
    
    def __init__(self,MyTest):
        # ---------------- [QA] Modify these variables for your test case! ----------------------------------       
        self.Feature= currentdir.split('\\')[-1]
        self.CaseTitle= '[UI] User auto play game'
        self.CaseNote= ''
        self.TestCaseID= '150611'
        # (CaseNote: can be blank ''. Use it to leave some notes or description of this case which will be displayed in test report)
        # ----------------------------------------------------------------------------------------------
        self.TestResult = 'FAIL'
        self.CurrentFile = os.path.basename(__file__)
        BaseCase.TestInfo = MyTest
        self.res_log = ''
        self.filename = tooldir+'\\code.jpg'  # I assume you have a way of picking unique filenames
        self.env = IniValue('setting', 'TestEnv.bc_ui_url')
        self.username = IniValue('setting', 'PlayGame.bc_acc')
        self.password = IniValue('setting', 'PlayGame.bc_pwd')
        # Determine which room for bet
        self.betting_room = IniValue('setting', 'PlayGame.bc_betting_room')
        self.betting_multiple = IniValue('setting', 'PlayGame.bc_betting_multiple')
        
    def setPreCondition(self):
        WriteLog('********Test Case: '+self.CurrentFile+'********')
        # [QA] Set up Pre-condition if any
        WriteLog('Set up Pre-condition of this case...')
        chrome_options = Options()
        chrome_options.add_argument('--allow-outdated-plugins')
        self.driver = webdriver.Chrome(tooldir+'\\chromedriver.exe', options=chrome_options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(self.TestInfo.normal_time)
        self.accept_next_alert = True
        self.CaseNote = WriteDebugLog(self.CaseNote, 'Success open browser.')
        self.driver.get(self.env)
        time.sleep(1)
        id_loginOpen = self.driver.find_element_by_id('loginOpen')
        id_loginOpen.click()
        username = self.driver.find_element_by_id('acc')
        username.click()
        username.send_keys(self.username)
#         username.send_keys(IniValue('setting', 'PlayGame.bc_acc'))
        password = self.driver.find_element_by_id('pwd')
        password.click()
        password.send_keys(self.password)
#         password.send_keys(IniValue('setting', 'PlayGame.bc_pwd'))
        self.driver.implicitly_wait(self.TestInfo.very_short_time)
        retry = 0
        while (retry < 10):
            try:
                self.driver.find_element_by_id("lgPassBtn").click()
                time.sleep(1)
                code = self.driver.find_element_by_id("code")
                code.clear()
                passCode = self.driver.find_element_by_id("passCode")
                imgdata = b64decode(passCode.get_attribute("src").split(',')[1])
                with open(self.filename, 'wb') as decode_response_data:
                    decode_response_data.write(imgdata)
                
                new_code, self.res_log = ApiCodeDetect('\n\t\t', self.res_log, self.filename)
                code.send_keys(new_code)
                time.sleep(0.5)
                next_btn = self.driver.find_element_by_id('loginIn')
                next_btn.click()
#                 id_loginNote = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'loginNote')),'Code fail')
#                 if(id_loginNote.text == '验证码错误'):
#                     continue
#                 else:
                id_informLg = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'informLg')),'BC_timeout (very_short_time)')
                if('登入成功' in id_informLg.text):
                    id_informLg.find_element_by_class_name('lg-close').click()
                    break
            except Exception as e:
                self.CaseNote = WriteDebugLog(self.CaseNote, 'Exception: on "Browser time out"... %s' % str(e))
            finally:
                retry += 1
        if(retry >= 10):
            self.CaseNote = WriteDebugLog(self.CaseNote, 'Retry browser fail.')
        
        
    def generateResults(self, collectOtherFile):
        # 1.[QA] if collectOtherFile == True, Collect temp results or other files if this cases needs...
        if (collectOtherFile):
            WriteLog('\t'+'Collect temp results or other files...')
            
            try:
                filepath1='TestDebug.log'
                if (os.path.exists(filepath1)):
                    shutil.copy(filepath1, self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature)
                    os.rename(self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature+"\\"+os.path.basename(filepath1), self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature+"\\"+self.CurrentFile+"_"+os.path.basename(filepath1))
            except Exception as e:
                WriteLog("Exception: on collectOtherFiles... %s" % str(e.with_traceback))
            WriteLog('...file '+ filepath1 + " copied")
         
        # 2. Run generateResults() of parent class
        else:
            super().generateResults()
        
        
    def teardown(self):
        # [QA] Tear down the environment if it needs
        WriteLog('\t'+'>> Tear Down of this case')
        self.driver.quit()
        WriteLog('\t\t'+'Driver closed')
        WriteLog('\t'+'<< End Tear Down')
        
        
    def run(self, copy_result):
        self.TestCaseStartTime = datetime.now()
        self.setPreCondition()
        print_tab = '\n\t\t'
        self.CaseNote = WriteDebugLog(self.CaseNote, '\t'+datetime.today().strftime("%m月%d日%H時%M分%S秒")+'>> Start to [UI] User auto play game'+'\n')
        #=======================================================================
        # [PlayGame]
        # game_provider = JDB_SLOT
        # game_name = 文房四宝
        # total_rounds = 10000
        # each_bet_dollar = 0.2
        #=======================================================================
        
#         game_type = IniValue('setting', 'PlayGame.game_type')
#         game_provider = IniValue('setting', 'PlayGame.game_provider')
#         game_name = IniValue('setting', 'PlayGame.game_name')
        game_type = 'poker'
        game_provider = 'JDB_POKER'
        game_name = '通比六牛'
        total_rounds = int(IniValue('setting', 'PlayGame.total_rounds'))
        
#         first_play = GraphicalLocator(tooldir+'\\asset\\bull6.png')
        
#         uWallect = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'uWallect')),'uWallect is not visible').text
        game_type_title = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, game_type+'pBtn')),'SLOT_title is not visible')
        hover = ActionChains(self.driver).move_to_element(game_type_title)
        hover.perform()
#         contentProvider = 'JDB_SLOT'
#         contentProvider = 'KA'
        gameBtn = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'a[href="game_list.html?type='+game_type+'&name='+game_provider+'"]')),'gameBtn is not visible')
        gameBtn.click()
        gameListArea_view = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'gameListArea')),'gameListArea_view is not visible')
        gameName_btns  = gameListArea_view.find_elements_by_css_selector('div[class="game-box"]')
        handle0 = self.driver.current_window_handle
#         playGame = '海盗船长'
#         playGame = '文房四宝'
        for gameName_btn in gameName_btns:
#             print(gameName_btn.text)
            if(game_name == gameName_btn.text):
                self.res_log = WriteDebugLog(self.res_log, print_tab+'- Into: '+gameName_btn.text)
                gm_btn = gameName_btn.find_element_by_css_selector('div[class="game-img"] div[class="game-btn"] a[onclick]')
                self.driver.execute_script('arguments[0].click();', gm_btn)
                self.driver.switch_to_window(self.driver.window_handles[1])
                self.driver.maximize_window()
                time.sleep(6)
                break
        
        align = Alignment(horizontal='left',vertical='top',wrap_text=False)
        gm_provider = game_provider.split('_')[0]
        excel_filename = tooldir+'\\'+gm_provider+'.xlsx'
        if(os.path.exists(excel_filename)):
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb.create_sheet(game_type+'_'+game_name, 0)
        else:
            wb = openpyxl.Workbook(excel_filename)
            wb.save(excel_filename)
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb['Sheet']
            ws.title = game_type+'_'+game_name
        
        print(wb.sheetnames)
        self.res_log = WriteDebugLog(self.res_log, print_tab+'- Excel Sheet: '+str(wb.sheetnames))
        ws.cell(row=1, column=1).value = '遊戲帳號'
        ws.cell(row=1, column=2).value = self.username
        ws.cell(row=1, column=3).value = '遊戲密碼'
        ws.cell(row=1, column=4).value = self.password
        ws.cell(row=1, column=5).value = '底注: '+str(self.betting_room).replace('room_','').replace('.png','')
        ws.cell(row=1, column=6).value = '倍數: '+str(self.betting_multiple).replace('bet_','').replace('.png','')
        
        ws.cell(row=2, column=1).value = '時間'
        ws.cell(row=2, column=2).value = '回合'
        ws.cell(row=2, column=3).value = '錢包'
#         ws.cell(row=3, column=3).value = uWallect
#         ws.cell(row=2, column=3).value = '期數'
#         ws.cell(row=2, column=4).value = '總分'
#         ws.cell(row=2, column=5).value = '投注額'
#         ws.cell(row=2, column=6).value = '贏分'
#         ws.cell(row=2, column=7).value = '輸贏'
#         ws.cell(row=2, column=8).value = '線數'
#         ws.cell(row=2, column=9).value = '比對餘額'
#         ws.cell(row=2, column=10).value = '有無FreeGame'

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["C"].width = 60
        ws.column_dimensions["D"].width = 40
#         ws.column_dimensions["D"].width = 20
#         ws.column_dimensions["F"].width = 15
#         ws.column_dimensions["I"].width = 10
#         ws.column_dimensions["J"].width = 15
        wb.save(excel_filename)
        
        # JDB Close Sound
        self.driver.switch_to.frame("gameframe")
        
        def checkSystemError(sleep_time):
            self.driver.switch_to_window(self.driver.current_window_handle)
            self.driver.implicitly_wait(sleep_time)
#             error_message = self.driver.find_element_by_css_selector('div[id="error"] div[class="error-inner"] div[class="error-content"] div[class="error-message"]')
            error_message = WebDriverWait(self.driver, sleep_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'error-message')),'error_message is not visible')
            error_button_list = error_message.find_elements_by_css_selector('div[class="call-action"] button[class="button"]')
            for error_button in error_button_list:
                if(error_button.text == '重读'):
                    error_button.click()
                    WriteLog('Error message click: '+error_button.text)
                    return True
                else:
                    WriteLog('Error message: '+error_button.text)
            return False
            
        # Join to game room
        click_width = 0
        click_heigh = 0
        timeout = self.TestInfo.normal_time
        while(click_width == 0 and click_heigh == 0 and timeout > 0):
            self.driver.switch_to_window(self.driver.current_window_handle)
            click_width, click_heigh = Compare_and_Click(self.driver, tooldir, tooldir+'\\asset\\'+self.betting_room, False, '', '')
            time.sleep(1)
            timeout -= 1
            if(timeout < 5):
                WriteLog('Popup Error Message')
                if(checkSystemError(self.TestInfo.very_short_time)):
                    break
                else:
                    timeout = 0
                    break
        if(timeout <= 0):
            WriteLog('Timeout Exception')
        
        action = ActionChains(self.driver)
        action.move_by_offset(click_width, click_heigh)
        action.click()
        action.perform()
        
        # Take a specific bet
        click_width = 0
        click_heigh = 0
        timeout = self.TestInfo.normal_time
        while(click_width == 0 and click_heigh == 0 and timeout > 0):
            self.driver.switch_to_window(self.driver.current_window_handle)
            click_width, click_heigh = Compare_and_Click(self.driver, tooldir, tooldir+'\\asset\\'+self.betting_multiple, False, '', '')
            print(click_width, click_heigh)
            time.sleep(1)
            timeout -= 1
            if(timeout < 5):
                WriteLog('Popup Error Message')
                if(checkSystemError(self.TestInfo.very_short_time)):
                    break
                else:
                    timeout = 0
                    break
        if(timeout <= 0):
            WriteLog('Timeout Exception')
        
        action = ActionChains(self.driver)
        action.move_by_offset(click_width, click_heigh)
        action.click()
        action.perform()
        
        # Bet until 1000 times
        i = 0
        while (i < total_rounds):
            # Detect continue button
            click_width = 0
            click_heigh = 0
            timeout = self.TestInfo.normal_time
            while(click_width == 0 and click_heigh == 0 and timeout > 0):
                self.driver.switch_to_window(self.driver.current_window_handle)
                click_width, click_heigh = Compare_and_Click(self.driver, tooldir, tooldir+'\\asset\\bull6_again.png', False, '', '')
                time.sleep(2)
                timeout -= 2
                if(timeout < 5):
                    WriteLog('Popup Error Message')
                    if(checkSystemError(self.TestInfo.very_short_time)):
                        break
                    else:
                        timeout = 0
                        break
            if(timeout <= 0):
                WriteLog('Timeout Exception')
                break
            
            # '時間'
            ws.cell(row=i+3, column=1).value = datetime.now()
            # '回合'
            ws.cell(row=i+3, column=2).value = str(i+1)
            ws.cell(row=i+3, column=1).alignment = align
            ws.row_dimensions[i+3].height = 30
            wb.save(excel_filename)
            WriteLog(print_tab+'回合: '+str(i+1))
            
            action = ActionChains(self.driver)
            action.move_by_offset(click_width, click_heigh)
            action.click()
            action.perform()
            
            # Take a specific bet
            click_width = 0
            click_heigh = 0
            timeout = self.TestInfo.normal_time
            while(click_width == 0 and click_heigh == 0 and timeout > 0):
                self.driver.switch_to_window(self.driver.current_window_handle)
                click_width, click_heigh = Compare_and_Click(self.driver, tooldir, tooldir+'\\asset\\'+self.betting_multiple, False, '', '')
                print(click_width, click_heigh)
                time.sleep(1)
                timeout -= 1
                if(timeout < 5):
                    WriteLog('Popup Error Message')
                    if(checkSystemError(self.TestInfo.very_short_time)):
                        break
                    else:
                        timeout = 0
                        break
            if(timeout <= 0):
                WriteLog('Timeout Exception')
                break
            
            action = ActionChains(self.driver)
            action.move_by_offset(click_width, click_heigh)
            action.click()
            action.perform()
            
            i += 1
        
        click_width = 0
        click_heigh = 0
        timeout = self.TestInfo.normal_time
        while(click_width == 0 and click_heigh == 0 and timeout > 0):
            self.driver.switch_to_window(self.driver.current_window_handle)
            click_width, click_heigh = Compare_and_Click(self.driver, tooldir, tooldir+'\\asset\\bull6_again.png', False, '', '')
            time.sleep(1)
            timeout -= 1
            if(timeout < 5):
                WriteLog('Popup Error Message')
                if(checkSystemError(self.TestInfo.very_short_time)):
                    break
                else:
                    timeout = 0
                    break
        if(timeout <= 0):
            WriteLog('Timeout Exception')
        
        # [QA] Be sure to save result "PASS" or "FAIL" to self.TestResult
        # ----------------------------------------------------------------------------------------------

        self.CaseNote = WriteDebugLog(self.CaseNote, '\n\t'+'#Test Result: ' + self.TestResult)
        self.teardown()

        self.TestCaseEndTime = datetime.now()
        self.TestCaseDuration = ((self.TestCaseEndTime - self.TestCaseStartTime).total_seconds())*1000000
        if (self.TestCaseDuration == 0):
            self.TestCaseDuration = ' less than 1'
        self.TestInfo.setTestRunEndTime()
        
        self.CaseNote = WriteDebugLog(self.CaseNote, '\n\t'+datetime.today().strftime("%m月%d日%H時%M分%S秒")+'<< End to [UI] User auto play game')
        # [QA] Set "collectOtherFile = True" if there is any file that should be collected
        TestCase.generateResults(self, copy_result)


# [QA] Be sure to test this case before it's merged to TestSuites
if __name__ == "__main__":
#             logging.debug()< logging.info()< logging.warning()< logging.error()< logging.critical()
    root_logger= logging.getLogger()
    root_logger.setLevel(logging.INFO) # or whatever
    handler = logging.FileHandler('TestDebug.log', 'w', 'utf-8') # or whatever
    handler.setFormatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s') # or whatever
    root_logger.addHandler(handler)
    
    WriteLog(" -------------------- Running a single Test Case --------------------")

    MyTest = TestRunning()
    MyTest.readTestConfig()
    if (MyTest.createTestRunFolder() == 'Error'):
        WriteLog("Test Tool will exit...")
        os._exit(1)

    thiscase = TestCase(MyTest)
    dir2 = MyTest.TestResultFolder + MyTest.TestRunFolder + thiscase.Feature
    if not os.path.exists(dir2):
        os.makedirs(dir2)

    thiscase.run(True)

    WriteLog('Test Case: ' + thiscase.CaseTitle + ', Test Result: ' + thiscase.TestResult)
    # Clear Test Tool Debug Log
    f = open("TestDebug.log", "w")
    f.flush()
    f.close()