#!/usr/bin/python
# -*- coding: UTF-8 -*-
import logging, sys, os, inspect, shutil, locale
from datetime import datetime
import time
from base64 import b64decode
from BaseTestCase import BaseCase
from O2oUtility import WriteLog, WriteDebugLog, IniValue, ApiCodeDetect
from TestRun import TestRunning

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import openpyxl
from openpyxl.styles import Alignment


locale.setlocale(locale.LC_CTYPE, 'chinese')
currentdir = os.path.dirname(os.path.abspath(inspect.getfile(inspect.currentframe())))
tooldir = os.path.dirname(os.path.dirname(currentdir))
if tooldir not in sys.path:
    sys.path.insert(0, tooldir)
iniFile = tooldir+'\\ini\\setting.ini'

class TestCase(BaseCase):
    
    def __init__(self,MyTest):
        # ---------------- [QA] Modify these variables for your test case! ----------------------------------       
        self.Feature= currentdir.split('\\')[-1]
        self.CaseTitle= '[UI] User auto play game'
        self.CaseNote= ''
        self.TestCaseID= '150611'
        # (CaseNote: can be blank ''. Use it to leave some notes or description of this case which will be displayed in test report)
        # ----------------------------------------------------------------------------------------------
        self.TestResult = 'FAIL'
        self.CurrentFile = os.path.basename(__file__)
        BaseCase.TestInfo = MyTest
        self.res_log = ''
        self.filename = tooldir+'\\code.jpg'  # I assume you have a way of picking unique filenames
        self.env = IniValue('setting', 'TestEnv.bc_ui_url')
        self.username = IniValue('setting', 'PlayGame.bc_acc')
        self.password = IniValue('setting', 'PlayGame.bc_pwd')
        
    def setPreCondition(self):
        WriteLog('********Test Case: '+self.CurrentFile+'********')
        # [QA] Set up Pre-condition if any
        WriteLog('Set up Pre-condition of this case...')
        chrome_options = Options()
        chrome_options.add_argument('--allow-outdated-plugins')
        self.driver = webdriver.Chrome(tooldir+'\\chromedriver.exe', options=chrome_options)
        self.driver.maximize_window()
        self.driver.implicitly_wait(self.TestInfo.normal_time)
        self.accept_next_alert = True
        self.CaseNote = WriteDebugLog(self.CaseNote, 'Success open browser.')
        self.driver.get(self.env)
        time.sleep(1)
        id_loginOpen = self.driver.find_element_by_id('loginOpen')
        id_loginOpen.click()
        username = self.driver.find_element_by_id('acc')
        username.click()
        username.send_keys(self.username)
#         username.send_keys(IniValue('setting', 'PlayGame.bc_acc'))
        password = self.driver.find_element_by_id('pwd')
        password.click()
        password.send_keys(self.password)
#         password.send_keys(IniValue('setting', 'PlayGame.bc_pwd'))
        self.driver.implicitly_wait(self.TestInfo.very_short_time)
        retry = 0
        while (retry < 10):
            try:
                self.driver.find_element_by_id("lgPassBtn").click()
                time.sleep(1)
                code = self.driver.find_element_by_id("code")
                code.clear()
                passCode = self.driver.find_element_by_id("passCode")
                imgdata = b64decode(passCode.get_attribute("src").split(',')[1])
                with open(self.filename, 'wb') as decode_response_data:
                    decode_response_data.write(imgdata)
                
                new_code, self.res_log = ApiCodeDetect('\n\t\t', self.res_log, self.filename)
                code.send_keys(new_code)
                time.sleep(0.5)
                next_btn = self.driver.find_element_by_id('loginIn')
                next_btn.click()
#                 id_loginNote = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'loginNote')),'Code fail')
#                 if(id_loginNote.text == '验证码错误'):
#                     continue
#                 else:
                id_informLg = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'informLg')),'BC_timeout (very_short_time)')
                if('登入成功' in id_informLg.text):
                    id_informLg.find_element_by_class_name('lg-close').click()
                    break
            except Exception as e:
                self.CaseNote = WriteDebugLog(self.CaseNote, 'Exception: on "Browser time out"... %s' % str(e))
            finally:
                retry += 1
        if(retry >= 10):
            self.CaseNote = WriteDebugLog(self.CaseNote, 'Retry browser fail.')
        
        
    def generateResults(self, collectOtherFile):
        # 1.[QA] if collectOtherFile == True, Collect temp results or other files if this cases needs...
        if (collectOtherFile):
            WriteLog('\t'+'Collect temp results or other files...')
            
            try:
                filepath1='TestDebug.log'
                if (os.path.exists(filepath1)):
                    shutil.copy(filepath1, self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature)
                    os.rename(self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature+"\\"+os.path.basename(filepath1), self.TestInfo.TestResultFolder+self.TestInfo.TestRunFolder+self.Feature+"\\"+self.CurrentFile+"_"+os.path.basename(filepath1))
            except Exception as e:
                WriteLog("Exception: on collectOtherFiles... %s" % str(e.with_traceback))
            WriteLog('...file '+ filepath1 + " copied")
         
        # 2. Run generateResults() of parent class
        else:
            super().generateResults()
        
        
    def teardown(self):
        # [QA] Tear down the environment if it needs
        WriteLog('\t'+'>> Tear Down of this case')
        self.driver.quit()
        WriteLog('\t\t'+'Driver closed')
        WriteLog('\t'+'<< End Tear Down')
        
        
    def run(self, copy_result):
        self.TestCaseStartTime = datetime.now()
        self.setPreCondition()
        result_list = []
        print_tab = '\n\t\t'
        self.CaseNote = WriteDebugLog(self.CaseNote, '\t'+datetime.today().strftime("%m月%d日%H時%M分%S秒")+'>> Start to [UI] User auto play game'+'\n')
        #=======================================================================
        # [PlayGame]
        # game_provider = JDB_SLOT
        # game_name = 文房四宝
        # total_rounds = 10000
        # each_bet_dollar = 0.2
        #=======================================================================
        
#         game_provider = IniValue('setting', 'PlayGame.game_provider')
#         game_name = IniValue('setting', 'PlayGame.game_name')
#         total_rounds = int(IniValue('setting', 'PlayGame.total_rounds'))
#         each_bet_dollar = IniValue('setting', 'PlayGame.each_bet_dollar')
        
        game_provider = 'JDB_SLOT'
        game_name = '文房四宝'
        total_rounds = 10000
        each_bet_dollar = '600'
        
        SLOT_title = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'slotpBtn')),'SLOT_title is not visible')
        hover = ActionChains(self.driver).move_to_element(SLOT_title)
        hover.perform()
#         contentProvider = 'JDB_SLOT'
#         contentProvider = 'KA'
        gameBtn = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'a[href="game_list.html?type=slot&name='+game_provider+'"]')),'gameBtn is not visible')
        gameBtn.click()
        gameListArea_view = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.ID, 'gameListArea')),'gameListArea_view is not visible')
        gameName_btns  = gameListArea_view.find_elements_by_css_selector('div[class="game-box"]')
        
#         playGame = '海盗船长'
#         playGame = '文房四宝'
        for gameName_btn in gameName_btns:
#             print(gameName_btn.text)
            if(game_name == gameName_btn.text):
                self.res_log = WriteDebugLog(self.res_log, print_tab+'- Into: '+gameName_btn.text)
                gm_btn = gameName_btn.find_element_by_css_selector('div[class="game-img"] div[class="game-btn"] a[onclick]')
                self.driver.execute_script('arguments[0].click();', gm_btn)
                self.driver.switch_to_window(self.driver.window_handles[1])
                self.driver.maximize_window()
                time.sleep(10)
                break
        
        align = Alignment(horizontal='left',vertical='top',wrap_text=False)
        excel_filename = tooldir+'\\'+game_provider+'.xlsx'
        if(os.path.exists(excel_filename)):
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb.create_sheet(game_name, 0)
        else:
            wb = openpyxl.Workbook(excel_filename)
            wb.save(excel_filename)
            wb = openpyxl.load_workbook(excel_filename)
            ws = wb['Sheet']
            ws.title = game_name
        
#         print(wb.sheetnames)
        self.res_log = WriteDebugLog(self.res_log, print_tab+'- Excel Sheet: '+str(wb.sheetnames))
        ws.cell(row=1, column=1).value = '遊戲帳號'
        ws.cell(row=1, column=2).value = self.username
        ws.cell(row=1, column=3).value = '遊戲密碼'
        ws.cell(row=1, column=4).value = self.password
        
        ws.cell(row=2, column=1).value = '時間'
        ws.cell(row=2, column=2).value = '回合'
        ws.cell(row=2, column=3).value = '期數'
        ws.cell(row=2, column=4).value = '總分'
        ws.cell(row=2, column=5).value = '投注額'
        ws.cell(row=2, column=6).value = '贏分'
        ws.cell(row=2, column=7).value = '輸贏'
        ws.cell(row=2, column=8).value = '線數'
        ws.cell(row=2, column=9).value = '比對餘額'
        ws.cell(row=2, column=10).value = '有無FreeGame'
        
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 20
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 15
        self.driver.switch_to.frame("gameframe")
        wb.save(excel_filename)
        
        # JDB Close Sound
        menuControl = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'menuControl')),'menuControl is not visible')
        menuButton = menuControl.find_element_by_class_name("menuButton")
        menuButton.click()
        soundBtn = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="menuControl"] div[class^="menuInfo"] ul[class="menuNav"] li button[class^="sound"]')),'button is not visible')
        soundBtn.click()
        
        # Turbo Spin
        turboIcon = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'turboIcon')),'turboIcon is not visible')
        turboIcon.click()
        
        # Check play number is difference
        def isFreeGame(total_before, total_after, bet, win):
            if(round(total_before-bet+win, 2) != total_after):
                return True
            else:
                return False
        
        # 投注金額
        bet = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'bet')),'bet is not visible').text
        
#         is_default_bet_dollar = WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.text_to_be_present_in_element((By.CLASS_NAME, 'bet'),'"'+each_bet_dollar+'"'))
        if(bet != each_bet_dollar):
            bet_dollar_btn = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'bet')),'bet_dollar_btn is not visible')
            bet_dollar_btn.click()
            betList_btns = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[class="layout_panelInnerBg"] div[class="betPanelInner"] div[class="betPanelInfo"] ul[class="betList"] li button')),'betList is not visible')
            for each_bet_btn in betList_btns:
                if(each_bet_btn.text.replace('\n','') == each_bet_dollar):
                    each_bet_btn.click()
                    self.res_log = WriteDebugLog(self.res_log, print_tab+'- Success change bet: '+ each_bet_btn.text.replace('\n',''))
                    bet = each_bet_dollar
                    break
        else:
            self.res_log = WriteDebugLog(self.res_log, print_tab+'- Success no change bet')
        
        # 線數
        lines = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="gameCreditPanelInner"] div[class="lines"] span[class^="creditColumn"]')),'lines is not visible')
        
        result_list.append('PASS')
        
        i = 0
        playerNumber_before = ''
        credit_before = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="gameCreditPanelInner"] div[class="credit"] span[class^="creditColumn"]')),'credit is not visible').text
        credit_before_float = float(credit_before.replace('¥','').replace(' ','').replace('\n','').replace(',',''))
        while(i < total_rounds):
            
            win = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="gameCreditPanelInner"] div[class="win"] span[class^="creditColumn"]')),'win is not visible')
            cal_win = float((win.text).replace('¥','').replace('\n','').replace(' ','').replace(',',''))
            cal_loss = float(bet.replace('\n','').replace(' ','').replace(',',''))
#             win_loss = round(cal_win-cal_loss, 2)
            if(i == 0):
                # '時間'
                ws.cell(row=i+3, column=1).value = datetime.now()
                ws.cell(row=i+3, column=1).alignment = align
                # '回合'
                ws.cell(row=i+3, column=2).value = str(i)
                # '期數'
                gameseqInner = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'gameseqInner')),'gameseqInner is not visible')
                ws.cell(row=i+3, column=3).value = gameseqInner.text
                # '總分'
                ws.cell(row=i+3, column=4).value = credit_before
                print('回合: '+str(i)+', 期數: '+gameseqInner.text+', 總分: '+credit_before)
                i += 1
                wb.save(excel_filename)
            else:
                try:
                    WebDriverWait(self.driver,self.TestInfo.very_short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[id="jdbGameContainer"][class^="basegame"]')),'basegame is not visible')
                    
                except TimeoutException as e:
                    if(str(e).replace('\n','') == 'Message: basegame is not visible'):
                        # Free game
                        WriteLog(print_tab+'This is FreeGame')
                        ws.cell(row=i+2, column=10).value = 'FreeGame'
                        ws.cell(row=i+2, column=10).alignment = align
    #                     self.res_log = WriteDebugLog(self.res_log, print_tab+'- This is FreeGame')
                
                try:
                    spinButton = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[class="spinButton layout_Button"]')),'spinButton is not visible')
                    spinButton.click()
                    time.sleep(0.5)
                    spinButton.click()
    #                 spinButton_spining = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[class="spinButton layout_Button isSpining"]')),'spinButton isSpining is not visible')
    #                 spinButton_spining.click()
                    
                    WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'button[class="spinButton layout_Button"]')),'spinButton is not visible')
                    playerNumber_after = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'gameseqInner')),'gameseqInner is not visible').text
                    time.sleep(1)
                    
                    if(playerNumber_before != playerNumber_after):
                        
                        playerNumber_before = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'gameseqInner')),'gameseqInner is not visible').text
                        
                        # 期數
                        gameseqInner = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'gameseqInner')),'gameseqInner is not visible')
                        
                        # 贏分
    #                     self.driver.implicitly_wait(30)
    #                     win = self.driver.find_element_by_css_selector('div[class^="gameCreditPanelInner"] div[class="win"] span[class^="creditColumn"]')
                        win = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="gameCreditPanelInner"] div[class="win"] span[class^="creditColumn"]')),'win is not visible')
                        
                        while(True):
                            win1 = win.text
                            time.sleep(0.5)
                            win2 = win.text
                            if(win1 == win2):
                                break
                        
                        # 總分
                        credit = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class^="gameCreditPanelInner"] div[class="credit"] span[class^="creditColumn"]')),'credit is not visible')
                        credit_after_float = float(credit.text.replace('¥','').replace(' ','').replace('\n','').replace(',',''))
                        
                        # '時間'
                        ws.cell(row=i+3, column=1).value = datetime.now()
                        # '回合'
                        ws.cell(row=i+3, column=2).value = str(i)
                        # '期數'
                        ws.cell(row=i+3, column=3).value = gameseqInner.text
                        # '總分'
                        ws.cell(row=i+3, column=4).value = credit.text
                        # '投注額'
                        ws.cell(row=i+3, column=5).value = bet
                        # '贏分'
                        ws.cell(row=i+3, column=6).value = win.text
                        
                        # '輸贏'
                        this_game_win = float((win.text).replace('¥','').replace('\n','').replace(' ','').replace(',',''))
                        this_game_loss = float(bet.replace('\n','').replace(' ','').replace(',',''))
                        win_loss = round(this_game_win-this_game_loss, 2)
                        ws.cell(row=i+3, column=7).value = str(win_loss)
                        
                        # '線數'
                        ws.cell(row=i+3, column=8).value = lines.text
                        
                        # '總分-投注+贏分=輸贏'
#                         print(credit_before_float)
#                         print(cal_win)
#                         print(cal_loss)
#                         print(round(credit_before_float+cal_win-cal_loss, 2))
#                         print(round(credit_after_float, 2))
#                         print(round(credit_before_float+cal_win-cal_loss, 2) == round(credit_after_float, 2))
                        compare_money = str(round(credit_before_float+cal_win-cal_loss, 2) == round(credit_after_float, 2))
                        if(compare_money == 'True'):
                            ws.cell(row=i+3, column=9).value = compare_money
                        else:
                            ws.cell(row=i+3, column=9).value = str(round(credit_before_float+cal_win-cal_loss, 2))+', '+str(round(credit_after_float, 2))
                        credit_before_float = credit_after_float
                        ws.cell(row=i+3, column=1).alignment = align
                        
                        wb.save(excel_filename)
    #                     self.res_log = WriteDebugLog(self.res_log, print_tab+'回合: '+str(i)+', 期數: '+gameseqInner.text+', 總分: '+credit.text+', 線數: '+lines.text+', 贏分: '+win.text+', 投注: '+bet+', 輸贏: '+str(win_loss))
                        WriteLog(print_tab+'回合: '+str(i)+', 期數: '+gameseqInner.text+', 總分: '+credit.text+', 投注: '+bet+', 贏分: '+win.text+ ', 輸贏: '+str(win_loss)+', 線數: '+lines.text+', 比對餘額: '+compare_money)
                        
                        i += 1
                    else:
                        time.sleep(0.5)
                
                except Exception as e:
                    self.res_log = WriteDebugLog(self.res_log, print_tab+'While exception: '+ str(e))
                    result_list.append('FAIL')
                    break
        
        result_list.append('PASS')
        self.res_log = WriteDebugLog(self.res_log, print_tab+'- Play round: '+ str(i))
        
        if('FAIL' in result_list or not result_list):
            self.TestResult = 'FAIL'
        else:
            self.TestResult = 'PASS'
        
#             else:
#                 print('xxx')
#                 freegame = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[id="jdbGameContainer"][class^="freegame"]')),'freegame is not visible')
#                 print('yyy')
#                 if(freegame):
#                     print('This is Freegame')
#                 spinButton = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CLASS_NAME, 'spinButton')),'spinButton is not visible')
#                 spinButton.click()
                
#         # Auto Run 999 times
#         autoControl = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[class="autoControl"]')),'autoControl is not visible')
#         autoButton_btn = autoControl.find_element_by_class_name("autoButton")
#         autoButton_btn.click()
# #                 autoButton_btn = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'autoButton')),'auto_btn is not visible')
# #                 autoButton_btn.click()
#         time.sleep(1)
#         autoLists = autoControl.find_elements_by_css_selector('div[class="autoTimesInner"] div[class="autoInfo"] ul[class="autoList"] li button')
# #                 autoLists = WebDriverWait(self.driver,self.TestInfo.short_time).until(EC.visibility_of_all_elements_located((By.CSS_SELECTOR, 'div[class="autoControl"] div[class="autoTimesInner"] div[class="autoInfo"] ul[class="autoList"] li button[class]')),'autoList is not visible')
#         for autoList in autoLists:
#             print(autoList.text)
#             if(autoList.text == '999'):
#                 autoList.click()
#                 time.sleep(20)
#                 break
#         print('AAAAAAAAAAAA')

        # [QA] Be sure to save result "PASS" or "FAIL" to self.TestResult
        # ----------------------------------------------------------------------------------------------

        self.CaseNote = WriteDebugLog(self.CaseNote, '\n\t'+'#Test Result: ' + self.TestResult)
        self.teardown()

        self.TestCaseEndTime = datetime.now()
        self.TestCaseDuration = ((self.TestCaseEndTime - self.TestCaseStartTime).total_seconds())*1000000
        if (self.TestCaseDuration == 0):
            self.TestCaseDuration = ' less than 1'
        self.TestInfo.setTestRunEndTime()
        
        self.CaseNote = WriteDebugLog(self.CaseNote, '\n\t'+datetime.today().strftime("%m月%d日%H時%M分%S秒")+'<< End to [UI] User auto play game')
        # [QA] Set "collectOtherFile = True" if there is any file that should be collected
        TestCase.generateResults(self, copy_result)


# [QA] Be sure to test this case before it's merged to TestSuites
if __name__ == "__main__":
#             logging.debug()< logging.info()< logging.warning()< logging.error()< logging.critical()
    root_logger= logging.getLogger()
    root_logger.setLevel(logging.INFO) # or whatever
    handler = logging.FileHandler('TestDebug.log', 'w', 'utf-8') # or whatever
    handler.setFormatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s') # or whatever
    root_logger.addHandler(handler)
    
    WriteLog(" -------------------- Running a single Test Case --------------------")

    MyTest = TestRunning()
    MyTest.readTestConfig()
    if (MyTest.createTestRunFolder() == 'Error'):
        WriteLog("Test Tool will exit...")
        os._exit(1)

    thiscase = TestCase(MyTest)
    dir2 = MyTest.TestResultFolder + MyTest.TestRunFolder + thiscase.Feature
    if not os.path.exists(dir2):
        os.makedirs(dir2)

    thiscase.run(True)

    WriteLog('Test Case: ' + thiscase.CaseTitle + ', Test Result: ' + thiscase.TestResult)
    # Clear Test Tool Debug Log
    f = open("TestDebug.log", "w")
    f.flush()
    f.close()